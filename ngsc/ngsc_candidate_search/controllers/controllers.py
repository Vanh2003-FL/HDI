# -*- coding: utf-8 -*-
import io
import json

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from odoo import http
from odoo.http import request


def parse_data(active_ids, candidate_search_id):
    # Xử lý active_ids
    active_ids = [int(id) for id in active_ids.split(',')] if active_ids else []
    # Xử lý candidate_search_id
    candidate_search_id = int(candidate_search_id) if candidate_search_id else False
    # Lấy bản ghi từ active_ids
    records = request.env['ngsc.candidate.search.result'].browse(active_ids).exists()
    candidate_search = request.env['ngsc.candidate.search'].browse(candidate_search_id).exists()

    # Thông tin chung employee
    emps = []
    for record in records:
        emps.append({
            'name': record.hr_employee_id.name,
            'email': record.work_email,
            'barcode': record.barcode,
            'job_title': record.job_title,
            'area_name': record.en_area_id.name,
            'block_name': record.en_block_id.name,
            'department': record.department_id.name,
            'department_name': record.en_department_id.name,
        })

    # Tạo điều kiện sắp xếp skill
    sort_cond = []
    for record in candidate_search.line_ids:
        sort_cond.append({
            'skill_id': record.skill_id,
            'level_id': record.level_id,
            'priority': record.priority,
            'weight': record.weight
        })

    sort_cond.sort(key=lambda x: (-x['priority'], -x['weight']))

    # Lấy danh sách tất cả skill của tất cả employee
    # employee_skills = request.env['hr.employee.skills'].search(
    #     [('hr_employee_id', 'in', [record.hr_employee_id.id for record in records])])

    # Lấy danh sách kỹ năng từ ngsc.candidate.search.result.skill
    employee_skills = http.request.env['ngsc.candidate.search.result.skill'].search([
        ('candidate_search_result_id', 'in', records.ids)
    ])

    all_skills = employee_skills.mapped('emp_skill_id.skill_id')

    # Đếm số nhân viên có mỗi kỹ năng để sắp xếp
    skill_employee_count = {}
    for skill in all_skills:
        count = len(employee_skills.filtered(lambda es: es.emp_skill_id.skill_id == skill))
        skill_employee_count[skill.id] = count

    # Sắp xếp all_skills theo:
    # 1. Có trong sort_cond (priority = True trước)
    # 2. Weight giảm dần
    # 3. Số nhân viên có kỹ năng (giảm dần)
    def skill_sort_key(skill):
        # Tìm skill trong sort_cond
        for idx, cond in enumerate(sort_cond):
            if cond['skill_id'].id == skill.id:
                return (
                    0,  # Có trong sort_cond
                    -cond['priority'],  # Priority = True xếp trước
                    -cond['weight'],  # Weight giảm dần
                    -skill_employee_count[skill.id]  # Số nhân viên giảm dần
                )
        return (
            1,  # Không có trong sort_cond
            0,  # Priority không áp dụng
            0,  # Weight không áp dụng
            -skill_employee_count[skill.id]  # Số nhân viên giảm dần
        )

    sorted_skills = all_skills.sorted(key=skill_sort_key)

    # Lấy thứ tự hr_employee_id từ records
    employee_order = [record.hr_employee_id.id for record in records]

    # Đếm số kỹ năng mà priority = True
    count_priority_true = len(list(filter(lambda x: x.get('priority') is True, sort_cond)))
    # Tạo bộ đếm để note số kỹ năng mà priority
    number_priority = 0

    # Tạo mảng 2 chiều, phần tử rỗng thì để None
    skill_matrix = []
    for skill in sorted_skills:

        if number_priority < count_priority_true:
            row = [1, skill]  # Cột đầu là đánh dấu kỹ năng có/ko priority, cột sau là skill
            number_priority += 1
        else:
            row = [0, skill]
            number_priority += 1

        for emp_id in employee_order:
            # Tìm skill tương ứng với employee_id và skill_id
            skill_record = employee_skills.filtered(
                lambda es: es.hr_employee_id.id == emp_id and es.emp_skill_id.skill_id == skill)
            row.append(skill_record if skill_record else None)
        skill_matrix.append(row)

    print(skill_matrix)

    return emps, skill_matrix


class CandidateSearchController(http.Controller):
    @http.route('/candidate/compare', type='http', auth='user', methods=['GET'])
    def compare_candidates(self, active_ids=None, candidate_search_id=None, **kwargs):
        emps, skill_matrix = parse_data(active_ids, candidate_search_id)

        # Chuẩn bị dữ liệu cho template
        values = {
            'records': emps,
            'skill_matrix': skill_matrix,
            'candidate_search_id': candidate_search_id,
            'title': "So sánh nhân sự",
            'full_width': True
        }
        return request.render('ngsc_candidate_search.compare_candidates', values)

    @http.route('/candidate/export_excel', type='http', auth='user', methods=['GET'], csrf=False)
    def export_excel(self, **kwargs):
        # Lấy active_ids và candidate_search_id từ query parameters
        active_ids = kwargs.get('active_ids', '')
        candidate_search_id = kwargs.get('candidate_search_id', None)

        emps, skill_matrix = parse_data(active_ids, candidate_search_id)

        # Tạo file Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"

        header_fill = PatternFill(start_color='71639e', end_color='71639e', fill_type='solid')  # Xanh lá nhạt
        headline_fill = PatternFill(start_color='aaa2c5', end_color='aaa2c5', fill_type='solid')  # Vàng nhạt
        thin_border = Border(left=Side(style=BORDER_THIN),
                             right=Side(style=BORDER_THIN),
                             top=Side(style=BORDER_THIN),
                             bottom=Side(style=BORDER_THIN))
        # Hàng header (Tên - Barcode, Job Title)
        ws.cell(row=1, column=1).value = ''
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).border = thin_border

        for col, emp in enumerate(emps, 2):
            ws.cell(row=1, column=col).value = f"{emp['name']} - {emp['barcode']}\n{emp['job_title']}"
            ws.cell(row=1, column=col).font = Font(bold=True, color='ffffff')
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', wrap_text=True)
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).border = thin_border

        # Phần Thông tin nhân sự
        row_idx = 2
        ws.cell(row=row_idx, column=1).value = "Thông tin nhân sự"
        ws.cell(row=row_idx, column=1).border = thin_border
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(emps) + 1)
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left')
        ws.cell(row=row_idx, column=1).fill = headline_fill
        row_idx += 1

        # Các hàng thông tin nhân sự
        personal_info = [
            ('Khu vực', 'area_name'),
            ('Khối', 'block_name'),
            ('Trung tâm/Ban', 'department'),
            ('Phòng', 'department_name'),
            ('Email', 'email')
        ]
        for label, field in personal_info:
            ws.cell(row=row_idx, column=1).value = label
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left')  # Căn trái cho cột đầu tiên
            ws.cell(row=row_idx, column=1).border = thin_border
            for col, emp in enumerate(emps, 2):
                ws.cell(row=row_idx, column=col).value = emp[field]
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

        # Phần Thông tin kỹ năng
        ws.cell(row=row_idx, column=1).value = "Thông tin kỹ năng"
        ws.cell(row=row_idx, column=1).border = thin_border
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(emps) + 1)
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left')
        ws.cell(row=row_idx, column=1).fill = headline_fill
        row_idx += 1

        # Lặp qua skill_matrix
        for skill_idx, row in enumerate(skill_matrix):
            # Hàng điểm đánh giá
            ws.cell(row=row_idx, column=1).value = row[0].name if row[0].name else ''
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left')  # Căn trái cho cột đầu tiên
            ws.cell(row=row_idx, column=1).border = thin_border
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            for col, col_data in enumerate(row[1:], 2):
                ws.cell(row=row_idx, column=col).value = col_data.emp_skill_id.current_level_id.name if col_data else ''
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=col).border = thin_border
                ws.cell(row=row_idx, column=1).font = Font(bold=True)
            row_idx += 1

            # Hàng trọng số
            ws.cell(row=row_idx, column=1).value = 'Trọng số'
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left')  # Căn trái cho cột đầu tiên
            ws.cell(row=row_idx, column=1).border = thin_border
            for col, col_data in enumerate(row[1:], 2):
                value = ''
                if col_data:
                    value = str(col_data.score) + '%' if col_data.score else ''
                ws.cell(row=row_idx, column=col).value = value
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

        # Set chiều rộng cột
        ws.column_dimensions['A'].width = 20  # Cột nhãn (Khu vực, Tên, Điểm đánh giá)
        for col in range(2, len(emps) + 2):  # Các cột ứng viên
            ws.column_dimensions[get_column_letter(col)].width = 50  # Chiều rộng cho tên, barcode, job title, v.v.

        # Lưu file Excel vào buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Trả về file Excel
        return request.make_response(
            output.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', 'attachment; filename=candidates.xlsx')
            ]
        )
