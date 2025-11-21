from odoo import models, fields, api, _, exceptions
from odoo.tools.misc import format_date, DEFAULT_SERVER_DATE_FORMAT
from datetime import timedelta, datetime, time
from odoo.tools import config, date_utils, get_lang, html2plaintext
from dateutil.relativedelta import relativedelta
from decimal import localcontext, Decimal, ROUND_HALF_UP
from pytz import timezone, UTC
from openpyxl import load_workbook
import io
import pathlib
from odoo.exceptions import ValidationError
from openpyxl.utils import get_column_letter


class CrmAccountReport(models.AbstractModel):
    _name = "crm.account.report"
    _description = "Báo cáo tình trạng Lead"
    _inherit = "account.report"

    filter_date = None
    filter_all_entries = None
    filter_journals = None
    filter_analytic = None
    filter_unfold_all = None

    @api.model
    def _get_columns(self, options):
        columns_monthly = []

        if options.get('filter_records'):
            records = self.env['crm.lead'].browse(options.get('filter_records'))
        else:
            records = self.env['crm.lead'].search([])
        # records = self.env['crm.lead']
        # records |= self.env['x.crm.lead.invoice'].search([('lead_id.company_id', 'in', self.env.companies.ids), ('date', '!=', False), ('date', '>=', date_from), ('date', '<=', date_to)]).mapped('lead_id')
        # records |= self.env['x.crm.lead.payment'].search([('lead_id.company_id', 'in', self.env.companies.ids), ('date', '!=', False), ('date', '>=', date_from), ('date', '<=', date_to)]).mapped('lead_id')
        types = records.line_ids.mapped('type_id')
        invoice_colspan = 0
        dated_lst = records.x_lead_invoice.filtered_domain([('date', '!=', False)]).mapped('date')
        if dated_lst:
            date_from = datetime.combine(min(dated_lst), time.min)
            date_to = datetime.combine(max(dated_lst), time.max)

            for date_step in date_utils.date_range(date_from, date_to, relativedelta(months=1)):
                # columns_monthly.append({'pre-offset': 18 + len(columns_monthly) + len(types), 'name': f'Tháng {date_step.month}/{date_step.year} <br/>({max(date_step + relativedelta(day=1), date_from).strftime("%d/%m")} -> {min(date_step + relativedelta(months=1, day=1, days=-1), date_to).strftime("%d/%m")})', 'style': 'background-color:#798AAF;text-align:center; white-space:nowrap; border:1px solid #000000'})
                columns_monthly.append({'pre-offset': 14 + 19 + invoice_colspan + len(types), 'name': f'Tháng {date_step.month}/{date_step.year}', 'style': 'background-color:#798AAF;text-align:center; white-space:nowrap; border:1px solid #000000'})
                invoice_colspan += 1
        else:
            columns_monthly.append({'pre-offset': 14 + 19 + invoice_colspan + len(types), 'name': f'', 'style': 'background-color:#798AAF;text-align:center; white-space:nowrap; border:1px solid #000000'})
            invoice_colspan += 1
        payment_colspan = 0
        dated_lst = records.x_lead_payment.filtered_domain([('date', '!=', False)]).mapped('date')
        if dated_lst:
            date_from = datetime.combine(min(dated_lst), time.min)
            date_to = datetime.combine(max(dated_lst), time.max)
            for date_step in date_utils.date_range(date_from, date_to, relativedelta(months=1)):
                # columns_monthly.append({'pre-offset': 19 + len(columns_monthly) + len(types), 'name': f'Tháng {date_step.month}/{date_step.year} <br/>({max(date_step + relativedelta(day=1), date_from).strftime("%d/%m")} -> {min(date_step + relativedelta(months=1, day=1, days=-1), date_to).strftime("%d/%m")})', 'style': 'background-color:#31869B;text-align:center; white-space:nowrap; border:1px solid #000000'})
                columns_monthly.append({'pre-offset': 15 + 19 + invoice_colspan + payment_colspan + len(types), 'name': f'Tháng {date_step.month}/{date_step.year}', 'style': 'background-color:#31869B;text-align:center; white-space:nowrap; border:1px solid #000000'})
                payment_colspan += 1
        else:
            columns_monthly.append({'pre-offset': 15 + 19 + invoice_colspan + payment_colspan + len(types), 'name': f'', 'style': 'background-color:#31869B;text-align:center; white-space:nowrap; border:1px solid #000000'})
            payment_colspan += 1
        columns_names = [
            {'name': 'STT', 'rowspan': 2, 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;border:1px solid #000000'},
            {'name': 'Ngày khởi tạo cơ hội KD', 'rowspan': 2, 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;border:1px solid #000000'},
            {'name': 'Mã/ID Lead', 'rowspan': 2, 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;border:1px solid #000000'},
            {'name': 'Số hợp đồng', 'rowspan': 2, 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;border:1px solid #000000'},
            {'name': 'Khách hàng', 'style': 'min-width:200px;background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Phân khúc', 'style': 'min-width:200px;background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Tên cơ hội/ dự án', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Mức độ ưu tiên', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': '% KT', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Sản phẩm', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Đơn vị bán', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Hãng', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Đối thủ cạnh tranh', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Mô hình vào thầu', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Vai trò của NGS', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Tổng giá trị dự án<br/>(Bao gồm cả VAT)', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Giá trị hợp đồng NGSC (Bao gồm cả VAT)', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2}, ]

        for type in types.sorted('id'):
            columns_names += [
                {'name': type.name, 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2}
            ]
        columns_names += [
            {'name': 'TSLN<br/>(Dự kiến)', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Lợi nhuận<br/>(Dự kiến)', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Ngày duyệt dự án ngân sách', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Ngày mở thầu', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Ngày đóng thầu', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Ngày ký HĐ', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Ngày kickoff (Khởi động/Triển khai) dự án', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Ngày kết thúc dự án', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Trạng thái', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'SA', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Đội tư vấn', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Đội sản xuất', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Nhân viên kinh doanh', 'rowspan': 2, 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;border:1px solid #000000'},
            {'name': 'Đội kinh<br/>doanh', 'rowspan': 2, 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;border:1px solid #000000'},
            {'name': 'Doanh thu xuất hoá đơn<br/>dự kiến', 'style': 'background-color:#798AAF;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Xuất hóa đơn', 'style': 'background-color:#798AAF;text-align:center; white-space:nowrap;', 'colspan': invoice_colspan},
            {'name': 'Tổng dòng tiền dự kiến', 'style': 'background-color:#31869B;text-align:center; white-space:nowrap;', 'rowspan': 2},
            {'name': 'Dòng tiền', 'style': 'background-color:#31869B;text-align:center; white-space:nowrap;', 'colspan': payment_colspan},
        ]

        return [columns_names, columns_monthly]

    @api.model
    def _get_report_name(self):
        return self._description

    def _get_reports_buttons(self, options):
        return [
            # {'name': _('PDF'), 'sequence': 1, 'action': 'print_pdf', 'file_export_type': _('PDF')},
            {'name': _('XLSX'), 'sequence': 2, 'action': 'print_xlsx', 'file_export_type': _('XLSX')},
        ]

    def _get_options(self, previous_options=None):
        res = super()._get_options(previous_options=previous_options)
        filter_records = previous_options and previous_options.get('filter_records') or []
        if self._context.get('filter_records') or filter_records:
            res['filter_records'] = self._context.get('filter_records') or filter_records
        if self._context.get('clear_filter_records'):
            res['filter_records'] = False
        return res

    @api.model
    def _get_lines(self, options, line_id=None):
        lines = []
        lg = self.env['res.lang']._lang_get(self.env.user.lang) or get_lang(self.env)
        if options.get('filter_records'):
            records = self.env['crm.lead'].browse(options.get('filter_records'))
        else:
            records = self.env['crm.lead'].search([])
        records = records.sudo()
        # records = self.env['crm.lead']
        # records |= self.env['x.crm.lead.invoice'].search([('lead_id.company_id', 'in', self.env.companies.ids), ('date', '!=', False), ('date', '>=', date_from), ('date', '<=', date_to)]).mapped('lead_id')
        # records |= self.env['x.crm.lead.payment'].search([('lead_id.company_id', 'in', self.env.companies.ids), ('date', '!=', False), ('date', '>=', date_from), ('date', '<=', date_to)]).mapped('lead_id')

        types = records.line_ids.mapped('type_id')

        invoice_datetime_from = False
        invoice_datetime_to = False

        dated_lst = records.x_lead_invoice.filtered_domain([('date', '!=', False)]).mapped('date')
        if dated_lst:
            invoice_datetime_from = datetime.combine(min(dated_lst), time.min)
            invoice_datetime_to = datetime.combine(max(dated_lst), time.max)

        payment_datetime_from = False
        payment_datetime_to = False
        dated_lst = records.x_lead_payment.filtered_domain([('date', '!=', False)]).mapped('date')
        if dated_lst:
            payment_datetime_from = datetime.combine(min(dated_lst), time.min)
            payment_datetime_to = datetime.combine(max(dated_lst), time.max)

        background = '#FFFFFF'
        with localcontext() as ctx:
            ctx.rounding = ROUND_HALF_UP
            for idx, record in enumerate(records, start=1):
                if background == '#FFFFFF':
                    background = '#D8DAE0'
                else:
                    background = '#FFFFFF'
                columns = [
                    {'name': UTC.localize(record.create_date).astimezone(timezone(self.env.user.tz or 'UTC')).replace(tzinfo=None).strftime(f'{lg.date_format} {lg.time_format}') if record.create_date else '', 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000'},
                    {'name': record.code or '', 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000'},
                    {'name': ', '.join(record.mapped('x_contract_ids.x_contract_code')), 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000'},
                    {'name': record.partner_id.name or '', 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:normal;border:1px solid #000000'},
                    {'name': record.partner_id.bfsi_id.name or '', 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:normal;border:1px solid #000000'},
                    {'name': record.name or '', 'style': f'background-color:{background};vertical-align:middle;text-align:left; white-space:normal;border:1px solid #000000'},
                    {'name': dict(record.fields_get(['priority'])['priority']['selection'])[record.priority] if record.priority else '', 'style': f'background-color:{background};vertical-align:middle;text-align:left; white-space:normal;border:1px solid #000000'},
                    {'name': f'{Decimal(record.probability).to_integral_value(rounding=ROUND_HALF_UP)}%', 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000'},
                    {'name': ', '.join([l.product_name for l in record.line_ids if l.product_name]), 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000'},
                    {'name': ', '.join(record.mapped('line_ids.supplier_id.name')), 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000'},
                    {'name': ', '.join(record.mapped('line_ids.partner_id.name')), 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000'},
                    {'name': ', '.join(record.mapped('opponent_ids.partner_id.name')), 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000'},
                    {'name': ', '.join(record.mapped('tender_model_id.name')), 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000'},
                    {'name': ', '.join(record.mapped('role_id.name')), 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000'},

                    {'name': self.format_value(record.company_currency._convert(record.expected_revenue, self.env.company.currency_id, self.env.company, record.date_deadline or fields.Datetime.now()), self.env.company.currency_id, True), 'style': f'background-color:{background};vertical-align:middle;text-align:right; white-space:nowrap;border:1px solid #000000'},
                    {'name': self.format_value(record.company_currency._convert(record.ngsd_revenue, self.env.company.currency_id, self.env.company, record.date_deadline or fields.Datetime.now()), self.env.company.currency_id, True), f'style': f'background-color:{background};vertical-align:middle;text-align:right; white-space:nowrap;border:1px solid #000000'},
                ]
                for type in types.sorted('id'):
                    columns += [
                        {'name': self.format_value(sum(record.company_currency._convert(line.ngsd_revenue, self.env.company.currency_id, self.env.company, record.date_deadline or fields.Datetime.now()) for line in record.line_ids.filtered(lambda x: x.type_id == type)), self.env.company.currency_id, True), 'style': f'background-color:{background};vertical-align:middle;text-align:right; white-space:nowrap;border:1px solid #000000'}
                    ]

                columns += [
                    {'name': f'{Decimal(record.x_expected_ratio * 100).to_integral_value(rounding=ROUND_HALF_UP)}%', 'style': f'background-color:{background};vertical-align:middle;text-align:right; white-space:nowrap;border:1px solid #000000'},
                    {'name': self.format_value(record.company_currency._convert(record.x_expected_margin, self.env.company.currency_id, self.env.company, record.date_deadline or fields.Datetime.now()), self.env.company.currency_id, True), 'style': f'background-color:{background};vertical-align:middle;text-align:right; white-space:nowrap;border:1px solid #000000'},

                    {'name': record.budget_approval_date.strftime(lg.date_format) if record.budget_approval_date else '', 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:normal;border:1px solid #000000'},
                    {'name': record.bid_opening_date.strftime(lg.date_format) if record.bid_opening_date else '', 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:normal;border:1px solid #000000'},
                    {'name': record.bid_closing_date.strftime(lg.date_format) if record.bid_closing_date else '', 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:normal;border:1px solid #000000'},
                    {'name': record.date_deadline.strftime(lg.date_format) if record.date_deadline else '', 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000'},
                    {'name': record.kickoff_planned_date.strftime(lg.date_format) if record.kickoff_planned_date else '', 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:normal;border:1px solid #000000'},
                    {'name': record.project_end_date.strftime(lg.date_format) if record.project_end_date else '', 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:normal;border:1px solid #000000'},
                    {'name': record.stage_id.name or '', 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000'},
                    {'name': ', '.join(record.mapped('solution_architect_uids.name')), 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:normal;border:1px solid #000000'},
                    {'name': ', '.join(record.mapped('x_consulting_team.name')), 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:normal;border:1px solid #000000'},
                    {'name': ', '.join(record.mapped('x_development_team.name')), 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:normal;border:1px solid #000000'},
                    {'name': record.user_id.name or '', 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000'},
                    {'name': record.team_id.name or '', 'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000'},

                ]

                invoice_columns = []
                payment_columns = []
                invoice_total = payment_total = 0
                if invoice_datetime_from and invoice_datetime_to:
                    for date_step in date_utils.date_range(invoice_datetime_from, invoice_datetime_to, relativedelta(months=1)):
                        compared_from = (max(date_step + relativedelta(day=1), invoice_datetime_from)).date()
                        compared_to = (min(date_step + relativedelta(months=1, day=1, days=-1), invoice_datetime_to)).date()
                        invoice_value = sum(line.currency_id._convert(line.amount, self.env.company.currency_id, self.env.company, line.date) for line in record.x_lead_invoice.filtered(lambda x: x.date and compared_from <= x.date <= compared_to))
                        invoice_columns += [
                            {'name': self.format_value(invoice_value, self.env.company.currency_id, True), 'style': f'background-color:#D8DAE0;vertical-align:middle;text-align:right; white-space:nowrap;border:1px solid #000000'}
                        ]
                        invoice_total += invoice_value
                else:
                    invoice_columns += [
                        {'name': '', 'style': f'background-color:#D8DAE0;vertical-align:middle;text-align:right; white-space:nowrap;border:1px solid #000000'}
                    ]
                if payment_datetime_from and payment_datetime_to:
                    for date_step in date_utils.date_range(payment_datetime_from, payment_datetime_to, relativedelta(months=1)):
                        compared_from = (max(date_step + relativedelta(day=1), payment_datetime_from)).date()
                        compared_to = (min(date_step + relativedelta(months=1, day=1, days=-1), payment_datetime_to)).date()
                        payment_value = sum(line.currency_id._convert(line.amount, self.env.company.currency_id, self.env.company, line.date) for line in record.x_lead_payment.filtered(lambda x: x.date and compared_from <= x.date <= compared_to))
                        payment_columns += [
                            {'name': self.format_value(payment_value, self.env.company.currency_id, True), 'style': f'background-color:#D8DAE0;vertical-align:middle;text-align:right; white-space:nowrap;border:1px solid #000000'}
                        ]
                        payment_total += payment_value
                else:
                    payment_columns += [
                        {'name': '', 'style': f'background-color:#D8DAE0;vertical-align:middle;text-align:right; white-space:nowrap;border:1px solid #000000'}
                    ]
                columns += [
                    {'name': self.format_value(invoice_total, self.env.company.currency_id, True), 'style': f'background-color:#798AAF;vertical-align:middle;text-align:right; white-space:nowrap;border:1px solid #000000'},
                ]
                columns += invoice_columns
                columns += [
                    {'name': self.format_value(payment_total, self.env.company.currency_id, True), 'style': f'background-color:#31869B;vertical-align:middle;text-align:right; white-space:nowrap;border:1px solid #000000'},
                ]
                columns += payment_columns

                lines += [{
                    'id': 'crm_lead_%s' % record.id,
                    'name': idx,
                    'style': f'background-color:{background};vertical-align:middle;text-align:center; white-space:nowrap;border:1px solid #000000',
                    'level': 1,
                    'columns': columns,
                }]

            ngsd_revenue = sum(record.company_currency._convert(record.ngsd_revenue, self.env.company.currency_id, self.env.company, record.date_deadline or fields.Datetime.now()) for record in records)

            columns = [
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': self.format_value(ngsd_revenue, self.env.company.currency_id, True), 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
            ]
            for type in types.sorted('id'):
                columns += [
                    {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'}
                ]

            columns += [
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
                {'name': '', 'style': 'background-color:#717A91;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
            ]

            invoice_columns = []
            payment_columns = []
            invoice_total = payment_total = 0
            if invoice_datetime_from and invoice_datetime_to:
                for date_step in date_utils.date_range(invoice_datetime_from, invoice_datetime_to, relativedelta(months=1)):
                    compared_from = (max(date_step + relativedelta(day=1), invoice_datetime_from)).date()
                    compared_to = (min(date_step + relativedelta(months=1, day=1, days=-1), invoice_datetime_to)).date()
                    invoice_value = sum(line.currency_id._convert(line.amount, self.env.company.currency_id, self.env.company, line.date) for line in records.mapped('x_lead_invoice').filtered(lambda x: x.date and compared_from <= x.date <= compared_to))
                    invoice_columns += [
                        {'name': self.format_value(invoice_value, self.env.company.currency_id, True), 'style': 'background-color:#798AAF;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'}
                    ]
                    invoice_total += invoice_value
            else:
                invoice_columns += [
                    {'name': '', 'style': 'background-color:#798AAF;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'}
                ]
            if payment_datetime_from and payment_datetime_to:
                for date_step in date_utils.date_range(payment_datetime_from, payment_datetime_to, relativedelta(months=1)):
                    compared_from = (max(date_step + relativedelta(day=1), payment_datetime_from)).date()
                    compared_to = (min(date_step + relativedelta(months=1, day=1, days=-1), payment_datetime_to)).date()
                    payment_value = sum(line.currency_id._convert(line.amount, self.env.company.currency_id, self.env.company, line.date) for line in records.mapped('x_lead_payment').filtered(lambda x: x.date and compared_from <= x.date <= compared_to))
                    payment_columns += [
                        {'name': self.format_value(payment_value, self.env.company.currency_id, True), 'style': 'background-color:#31869B;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'}
                    ]
                    payment_total += payment_value
            else:
                payment_columns += [
                    {'name': '', 'style': 'background-color:#31869B;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'}
                ]
            columns += [
                {'name': f'<span style="color:red">{self.format_value(invoice_total, self.env.company.currency_id, True)}</span>', 'style': 'background-color:#798AAF;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
            ]
            columns += invoice_columns
            columns += [
                {'name': f'<span style="color:red">{self.format_value(payment_total, self.env.company.currency_id, True)}</span>', 'style': 'background-color:#31869B;vertical-align:middle;text-align:right; white-space:nowrap;border:0px solid #000000'},
            ]
            columns += payment_columns

            lines += [{
                'id': 'total_crm_lead_%s' % self.env.user.id,
                'name': '',
                'class': 'total',
                'style': 'background-color:#717A91;vertical-align:middle;text-align:center; white-space:nowrap;border:0px solid #000000',
                'level': 1,
                'columns': columns,
            }]
        return lines


    def get_xlsx(self, options, response=None):
        output = io.BytesIO()
        template_name = 'lead_report.xlsx'
        template_file = pathlib.Path(__file__).parent.parent.joinpath('static', 'excel', template_name)
        try:
            workbook = load_workbook(template_file)
        except FileNotFoundError:
            raise ValidationError(_("Error: Could not find template {}").format(template_name))

        sheet = workbook['Mẫu báo cáo']

        # Set the first column width to 50
        # sheet.set_column(0, 0, 50)

        y_offset = 1
        headers, lines = self.with_context(no_format=True, print_mode=True, prefetch_fields=False)._get_table(options)

        max_header_row = 34

        max_row = 0
        for header in headers:
            x_offset = 1
            for column in header:
                colspan = column.get('colspan', 1)
                pre_offset = column.get('pre-offset', 0)
                if pre_offset:
                    x_offset = pre_offset
                x_offset += colspan
            max_row = max(max_row, x_offset)
        if max_row > max_header_row:
            for i in range(1, max_row - max_header_row):
                sheet.insert_cols(max_header_row+1)
                sheet.column_dimensions[get_column_letter(max_header_row+1)].width = 14
                sheet.cell(1, max_header_row+1)._style = sheet.cell(1, max_header_row)._style
                sheet.cell(2, max_header_row+1)._style = sheet.cell(2, max_header_row)._style
                sheet.cell(3, max_header_row+1)._style = sheet.cell(3, max_header_row)._style

        for header in headers:
        # Add headers.
            x_offset = 1
            for column in header:
                column_name_formated = html2plaintext(column.get('name', '').replace('<br/>', ' ').replace('&nbsp;', ' '))
                colspan = column.get('colspan', 1)
                rowspan = column.get('rowspan', 1)
                pre_offset = column.get('pre-offset', 0)
                if pre_offset:
                    x_offset = pre_offset

                sheet.cell(y_offset, x_offset, column_name_formated)
                if colspan != 1 or rowspan != 1:
                    sheet.merge_cells(start_row=y_offset, start_column=x_offset, end_row=y_offset + rowspan - 1, end_column=x_offset + colspan - 1)
                x_offset += colspan
            y_offset += 1

        if options.get('hierarchy'):
            lines = self._create_hierarchy(lines, options)
        if options.get('selected_column'):
            lines = self._sort_lines(lines, options)

        # Add lines.
        for y in range(0, len(lines)):
        #     level = lines[y].get('level')
        #     if lines[y].get('caret_options'):
        #         style = level_3_style
        #         col1_style = level_3_col1_style
        #     elif level == 0:
        #         y_offset += 1
        #         style = level_0_style
        #         col1_style = style
        #     elif level == 1:
        #         style = level_1_style
        #         col1_style = style
        #     elif level == 2:
        #         style = level_2_style
        #         col1_style = 'total' in lines[y].get('class', '').split(' ') and level_2_col1_total_style or level_2_col1_style
        #     elif level == 3:
        #         style = level_3_style
        #         col1_style = 'total' in lines[y].get('class', '').split(' ') and level_3_col1_total_style or level_3_col1_style
        #     else:
        #         style = default_style
        #         col1_style = default_col1_style

            # write the first column, with a specific style to manage the indentation
            cell_type, cell_value = self._get_cell_type_value(lines[y])
            # if cell_type == 'date':
            #     sheet.write_datetime(y + y_offset, 0, cell_value, date_default_col1_style)
            # else:
            sheet.cell(y + y_offset, 1, cell_value)
            sheet.cell(y + y_offset, 1)._style = sheet.cell(3, 1)._style
            # write all the remaining cells
            for x in range(1, len(lines[y]['columns']) + 1):
                cell_type, cell_value = self._get_cell_type_value(lines[y]['columns'][x - 1])
                # if cell_type == 'number':
                #     sheet.write_number(y + y_offset, x + lines[y].get('colspan', 1) - 1, cell_value, style)
                # elif cell_type == 'date':
                #     sheet.write_datetime(y + y_offset, x + lines[y].get('colspan', 1) - 1, cell_value, date_default_style)
                # else:
                sheet.cell(y + y_offset, x + lines[y].get('colspan', 1), cell_value)
                sheet.cell(y + y_offset, x + lines[y].get('colspan', 1))._style = sheet.cell(3, x + lines[y].get('colspan', 1))._style
        # sheet.autofit()
        workbook.save(output)
        output.seek(0)
        generated_file = output.read()
        output.close()

        return generated_file
