from odoo import models, fields, api, _, exceptions
from odoo.tools.misc import format_date, DEFAULT_SERVER_DATE_FORMAT
from datetime import timedelta, datetime, time
from odoo.tools import config, date_utils, get_lang, html2plaintext
from dateutil.relativedelta import relativedelta
from decimal import localcontext, Decimal, ROUND_HALF_UP

from openpyxl import load_workbook
import io
import pathlib
from odoo.exceptions import ValidationError
from openpyxl.utils import get_column_letter

import operator as py_operator

OPERATORS = {
    '<': py_operator.lt,
    '>': py_operator.gt,
    '<=': py_operator.le,
    '>=': py_operator.ge,
    '=': py_operator.eq,
    '!=': py_operator.ne
}


class KpiAccountReportWizard(models.TransientModel):
    _name = "kpi.account.report.wizard"
    _description = "Báo cáo thực hiện KPI"

    date_from = fields.Date(string='Từ ngày')
    date_to = fields.Date(string='Đến ngày')
    type = fields.Selection(string='Đối tượng', selection=[('team', 'Nhóm kinh doanh'), ('sale', 'Nhân viên kinh doanh')], default='team')
    sale_team_ids = fields.Many2many(string='Nhóm kinh doanh', comodel_name='crm.team')
    user_ids = fields.Many2many(string='Nhân viên kinh doanh', comodel_name='res.users')
    period = fields.Selection(string='Kỳ KPI', selection=[('month', 'Tháng'), ('quarter', 'Quý'), ('year', 'Năm')])

    technical = fields.Boolean(string='🐧', compute='_compute_technical', search='_search_technical')

    @api.depends('sale_team_ids', 'user_ids')
    def _compute_technical(self):
        for rec in self:
            technical = True
            for sale in rec.sale_team_ids:
                if self.env.user not in sale.user_ids:
                    technical = False
                    break
            for user in rec.user_ids:
                if self.env.user not in user.sale_team_id.user_ids:
                    technical = False
                    break
            rec.technical = technical

    def _search_technical(self, operator, value):
        if operator not in ('=', '!='):
            raise exceptions.UserError(_('Invalid domain operator %s', operator))
        ids = []
        for record in self.with_context(prefetch_fields=False).search([], order='id desc'):
            if OPERATORS[operator](record['technical'], value):
                ids.append(record.id)
        return [('id', 'in', ids)]

    def do(self):
        self = self.sudo()
        action = self.env.ref('account_reports.action_kpi_account_report').read()[0]
        action['target'] = 'main'
        action['context'] = {'model': 'kpi.account.report',
                             # 'date_from': self.date_from,
                             # 'date_to': self.date_to,
                             'type': self.type or 'team',
                             'sale_team_ids': self.sale_team_ids.ids,
                             'user_ids': self.user_ids.ids,
                             'period': self.period or 'month',
                             'forcely': True}
        return action


class KpiAccountReport(models.AbstractModel):
    _name = "kpi.account.report"
    _description = "Báo cáo thực hiện KPI"
    _inherit = "account.report"

    filter_date = {'mode': 'range', 'filter': 'this_year'}
    filter_all_entries = None
    filter_journals = None
    filter_analytic = None
    filter_unfold_all = None

    @api.model
    def _get_columns(self, options):
        columns_names = [
            {'name': 'Nhóm kinh doanh' if options.get('type') != 'sale' else 'Nhân viên kinh doanh', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;border:1px solid #000000'},
            {'name': 'Chu kỳ', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;border:1px solid #000000'},
            {'name': 'Từ ngày', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;border:1px solid #000000'},
            {'name': 'Đến ngày', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;border:1px solid #000000'},
            {'name': 'Chỉ tiêu doanh số', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;border:1px solid #000000'},
            {'name': 'Dự kiến doanh số', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;'},
            {'name': 'Thực tế doanh số', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;'},
            {'name': 'Chỉ tiêu dòng tiền', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;border:1px solid #000000'},
            {'name': 'Dự kiến dòng tiền', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;'},
            {'name': 'Thực tế dòng tiền', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;'},
            {'name': 'Chỉ tiêu doanh thu', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;border:1px solid #000000'},
            {'name': 'Dự kiến doanh thu', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;'},
            {'name': 'Thực tế doanh thu', 'style': 'background-color:#3F4C6A;color:white;text-align:center; white-space:nowrap;'},
        ]

        return [columns_names]

    @api.model
    def _get_report_name(self):
        return self._description

    def _get_reports_buttons(self, options):
        return [
            # {'name': _('PDF'), 'sequence': 1, 'action': 'print_pdf', 'file_export_type': _('PDF')},
            {'name': _('XLSX'), 'sequence': 2, 'action': 'print_xlsx', 'file_export_type': _('XLSX')},
        ]

    def _set_context(self, options):
        ctx = super()._set_context(options)
        # date_from = ctx.get('date_from')
        # date_to = ctx.get('date_to')
        type = ctx.get('type')
        sale_team_ids = ctx.get('sale_team_ids')
        user_ids = ctx.get('user_ids')
        period = ctx.get('period')
        forcely = ctx.get('forcely')
        if not type:
            wizard = self.env['kpi.account.report.wizard'].search([('create_uid', '=', self.env.user.id)], order='id desc', limit=1)
            # date_from = wizard.date_from
            # date_to = wizard.date_to
            type = wizard.type
            sale_team_ids = wizard.sale_team_ids.ids
            user_ids = wizard.user_ids.ids
            period = wizard.period

        # ctx['date_from'] = date_from
        # ctx['date_to'] = date_to
        ctx['type'] = type or 'team'
        ctx['sale_team_ids'] = sale_team_ids
        ctx['user_ids'] = user_ids
        ctx['period'] = period or 'month'
        ctx['forcely'] = forcely
        return ctx

    def _get_options(self, previous_options=None):
        res = super()._get_options(previous_options=previous_options)
        lst_key = ['type', 'sale_team_ids', 'user_ids', 'period', 'forcely']
        for k in lst_key:
            v = previous_options and previous_options.get(k) or []
            if self._context.get(k) or v:
                res[k] = self._context.get(k) or v
        return res

    @api.model
    def _get_lines(self, options, line_id=None):
        lines = []
        lg = self.env['res.lang']._lang_get(self.env.user.lang) or get_lang(self.env)
        date_from = datetime.combine(min(fields.Date.from_string(options['date']['date_from']), fields.Date.from_string(options['date']['date_to'])), time.min)
        date_to = datetime.combine(max(fields.Date.from_string(options['date']['date_from']), fields.Date.from_string(options['date']['date_to'])), time.max)
        ctx = self._context

        type = options.get('type')
        sale_team_ids = options.get('sale_team_ids')
        user_ids = options.get('user_ids')
        period = options.get('period')

        if type != 'sale':
            domain = [('type', '=', 'team'), ]
            if sale_team_ids:
                domain += [('sale_team_id', 'in', sale_team_ids)]
        else:
            domain = [('type', '=', 'sale')]
            if user_ids:
                domain += [('user_id', 'in', user_ids)]
        domain += ['|',
                   '&', ('date_from', '<=', date_from.date()), ('date_to', '>=', date_from.date()),
                   '&', ('date_from', '>=', date_from.date()), ('date_from', '<=', date_to.date()),
                   ]
        details = self.env['en.kpi.detail'].search(domain)
        month_step = 1
        if period == 'quarter':
            month_step = 3
        elif period == 'year':
            month_step = 12
        sequence = 1
        background = '#FFFFFF'
        for dately in date_utils.date_range(date_from + relativedelta(month=1, day=1), date_to + relativedelta(years=1, month=1, day=1, days=-1), relativedelta(months=month_step)):
            if dately > date_to: continue
            if dately + relativedelta(months=month_step) + relativedelta(days=-1) < date_from: continue
            if background == '#FFFFFF':
                background = '#D8DAE0'
            else:
                background = '#FFFFFF'
            line_date_from = max(dately, date_from)
            line_date_to = min(dately + relativedelta(months=month_step) + relativedelta(days=-1), date_to)
            # line_date_from = dately
            # line_date_to = dately + relativedelta(months=month_step) + relativedelta(days=-1)
            records = details.mapped('sale_team_id') if type != 'sale' else details.mapped('user_id')
            if period == 'quarter':
                period_name = f'Q{(line_date_from.month - 1) // 3 + 1} {line_date_from.year}'
            elif period == 'year':
                period_name = {line_date_from.year}
            else:
                period_name = f'T{line_date_from.month} {line_date_from.year}'
            for record in set(records):

                filtered_domain = []
                if type != 'sale':
                    filtered_domain = [('type', '=', 'team'), ]
                    if sale_team_ids:
                        filtered_domain += [('sale_team_id', 'in', record.ids)]
                else:
                    filtered_domain = [('type', '=', 'sale')]
                    if user_ids:
                        filtered_domain += [('user_id', 'in', record.ids)]
                filtered_domain += ['|',
                                    '&', ('date_from', '<=', line_date_from.date()), ('date_to', '>=', line_date_from.date()),
                                    '&', ('date_from', '>=', line_date_from.date()), ('date_from', '<=', line_date_to.date()),
                                    ]
                record_details = details.filtered_domain(filtered_domain)
                columns = [
                    {'name': period_name, 'style': f'background-color:{background};text-align:center;vertical-align:middle; white-space:nowrap;border:1px solid #000000;'},
                    {'name': line_date_from.strftime(lg.date_format), 'class': 'date', 'style': f'background-color:{background};text-align:center;vertical-align:middle; white-space:nowrap;border:1px solid #000000;'},
                    {'name': line_date_to.strftime(lg.date_format), 'class': 'date', 'style': f'background-color:{background};text-align:center;vertical-align:middle; white-space:nowrap;border:1px solid #000000;'},
                    {'name': self.format_value(sum(record_details.mapped('kpi_sales')), self.env.company.currency_id, True), 'style': f'background-color:{background};text-align:right;vertical-align:middle; white-space:nowrap;border:1px solid #000000;'},
                    {'name': self.format_value(sum(record_details.mapped('expected_sales')), self.env.company.currency_id, True), 'style': f'background-color:{background};text-align:right;vertical-align:middle; white-space:nowrap;border:1px solid #000000;'},
                    {'name': self.format_value(sum(record_details.mapped('sales')), self.env.company.currency_id, True), 'style': f'background-color:{background};text-align:right;vertical-align:middle; white-space:nowrap;border:1px solid #000000;'},
                    {'name': self.format_value(sum(record_details.mapped('kpi_revenue')), self.env.company.currency_id, True), 'style': f'background-color:{background};text-align:right;vertical-align:middle; white-space:nowrap;border:1px solid #000000;'},
                    {'name': self.format_value(sum(record_details.mapped('expected_revenue')), self.env.company.currency_id, True), 'style': f'background-color:{background};text-align:right;vertical-align:middle; white-space:nowrap;border:1px solid #000000;'},
                    {'name': self.format_value(sum(record_details.mapped('revenue')), self.env.company.currency_id, True), 'style': f'background-color:{background};text-align:right;vertical-align:middle; white-space:nowrap;border:1px solid #000000;'},
                    {'name': self.format_value(sum(record_details.mapped('kpi_invoiced')), self.env.company.currency_id, True), 'style': f'background-color:{background};text-align:right;vertical-align:middle; white-space:nowrap;border:1px solid #000000;'},
                    {'name': self.format_value(sum(record_details.mapped('expected_invoiced')), self.env.company.currency_id, True), 'style': f'background-color:{background};text-align:right;vertical-align:middle; white-space:nowrap;border:1px solid #000000;'},
                    {'name': self.format_value(sum(record_details.mapped('invoiced')), self.env.company.currency_id, True), 'style': f'background-color:{background};text-align:right;vertical-align:middle; white-space:nowrap;border:1px solid #000000;'},
                ]
                lines.append({
                    'id': f'kpi_{record.id}_{line_date_from.strftime("%m/%m/%Y")}',
                    'name': record.name,
                    'style': f'background-color:{background};text-align:center;vertical-align:middle; white-space:nowrap;border:1px solid #000000;',
                    'columns': columns,
                    'level': 1,
                })
            sequence += 1
        return lines

    def get_xlsx(self, options, response=None):
        output = io.BytesIO()
        template_name = 'kpi_report.xlsx'
        template_file = pathlib.Path(__file__).parent.parent.joinpath('static', 'excel', template_name)
        try:
            workbook = load_workbook(template_file)
        except FileNotFoundError:
            raise ValidationError(_("Error: Could not find template {}").format(template_name))

        sheet = workbook['Báo cáo KPI']

        # Set the first column width to 50
        # sheet.set_column(0, 0, 50)

        y_offset = 2
        headers, lines = self.with_context(no_format=True, print_mode=True, prefetch_fields=False)._get_table(options)

        max_init_header_col = 13
        max_header_col = 0
        for header in headers:
            x_offset = 1
            for column in header:
                colspan = column.get('colspan', 1)
                pre_offset = column.get('pre-offset', 0)
                if pre_offset:
                    x_offset = pre_offset
                x_offset += colspan
            max_header_col = max(max_header_col, x_offset)
        if max_header_col > max_init_header_col:
            for i in range(1, max_header_col - max_init_header_col):
                sheet.insert_cols(max_init_header_col+1)
                sheet.column_dimensions[get_column_letter(max_init_header_col+1)].width = 14
                sheet.cell(1, max_init_header_col+1)._style = sheet.cell(1, max_init_header_col)._style
                sheet.cell(2, max_init_header_col+1)._style = sheet.cell(2, max_init_header_col)._style
                sheet.cell(3, max_init_header_col+1)._style = sheet.cell(3, max_init_header_col)._style

        for header in headers:
        # Add headers.
            x_offset = 1
            for column in header:
                column_name_formated = html2plaintext(column.get('name', '').replace('<br/>', ' ').replace('&nbsp;', ' '))
                colspan = column.get('colspan', 1)
                rowspan = column.get('rowspan', 1)
                pre_offset = column.get('pre-offset', 0)
                if pre_offset:
                    x_offset = pre_offset

                sheet.cell(y_offset, x_offset, column_name_formated)
                if colspan != 1 or rowspan != 1:
                    sheet.merge_cells(start_row=y_offset, start_column=x_offset, end_row=y_offset + rowspan - 1, end_column=x_offset + colspan - 1)
                x_offset += colspan
            y_offset += 1

        if options.get('hierarchy'):
            lines = self._create_hierarchy(lines, options)
        if options.get('selected_column'):
            lines = self._sort_lines(lines, options)

        # Add lines.
        for y in range(0, len(lines)):
            # write the first column, with a specific style to manage the indentation
            cell_type, cell_value = self._get_cell_type_value(lines[y])
            sheet.cell(y + y_offset, 1, cell_value)
            sheet.cell(y + y_offset, 1)._style = sheet.cell(3, 1)._style
            # write all the remaining cells
            for x in range(1, len(lines[y]['columns']) + 1):
                cell_type, cell_value = self._get_cell_type_value(lines[y]['columns'][x - 1])
                sheet.cell(y + y_offset, x + lines[y].get('colspan', 1), cell_value)
                sheet.cell(y + y_offset, x + lines[y].get('colspan', 1))._style = sheet.cell(3, x + lines[y].get('colspan', 1))._style
        # sheet.autofit()
        workbook.save(output)
        output.seek(0)
        generated_file = output.read()
        output.close()

        return generated_file
