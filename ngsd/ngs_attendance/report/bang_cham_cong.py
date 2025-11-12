from collections import defaultdict

from dateutil.relativedelta import relativedelta

from odoo import models, _,fields
import pathlib
from odoo.exceptions import ValidationError
from openpyxl import load_workbook
from datetime import timedelta, datetime, date, time

from pytz import timezone

from openpyxl.styles import PatternFill

redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')


class AttendanceExportReportXlsx(models.AbstractModel):
    _name = 'report.ngs_attendance.report_attendance'
    _inherit = 'report.report_xlsx.abstract'

    def generate_xlsx_report(self, workbook, data, objs):
        template_name = 'bang_cham_cong.xlsx'
        template_file = pathlib.Path(__file__).parent.parent.joinpath('static', 'excel', template_name)
        try:
            wb = load_workbook(template_file)
        except FileNotFoundError:
            raise ValidationError(_("Error: Could not find template {}").format(template_name))

        sheet = wb['Bảng chấm công']
        if objs.export_option == 'month':
            # Xuất theo tháng
            return self._export_by_month(wb, sheet, objs)
        else:
            # Xuất theo tùy chọn khoảng thời gian
            if objs.date_from and objs.date_to and (objs.date_to - objs.date_from).days > 30:
                raise ValidationError("Khoảng thời gian xuất bảng chấm công không được quá 31 ngày!")
            return self._export_by_custom(wb, sheet, objs)

    def _export_by_month(self, wb, sheet, objs):
        row_to_write = 7
        name = 'Bảng chấm công tháng %s-%s' % (objs.month, objs.year)
        stt = 0
        sheet['B2'] = objs.month
        sheet['B3'] = objs.year
        sheet['N3'] = f'Tháng {objs.month}/{objs.year}'
        day_of_week_map = {0: 'T2', 1: 'T3', 2: 'T4', 3: 'T5', 4: 'T6', 5: 'T7', 6: 'CN'}
        day_in_month_col = 8
        day_start = datetime(int(objs.year), int(objs.month), 1)
        day_end = (datetime(int(objs.year), int(objs.month), 1) + relativedelta(months=1, days=-1))

        max_day = day_end.day
        for i in range(1, 32):
            if i <= max_day:
                sheet.cell(row_to_write-2, day_in_month_col + i, day_of_week_map.get(datetime(int(objs.year), int(objs.month), i).weekday()))
                sheet.cell(row_to_write-1, day_in_month_col + i, i)

        employees = objs.with_context(active_test=False).employee_ids
        if not employees:
            domain = ['|', ('active', '=', True),
                      '&',
                        ('active', '=', False),
                        '|',
                            '&',
                                '|', ('en_day_layoff_from', '<=', day_end), ('en_day_layoff_from', '=', False),
                                '|', ('en_day_layoff_to', '>=', day_start), ('en_day_layoff_to', '=', False),
                            ('departure_date', '>', day_start)
                      ]
            if objs.department_ids:
                domain += [('department_id','in',objs.department_ids.ids)]
            if objs.en_department_ids:
                domain += [('en_department_id','in',objs.en_department_ids.ids)]

            employees = self.env['hr.employee'].search(domain)
        check_hour = float(self.env['ir.config_parameter'].sudo().get_param('check_timesheet_before_checkout_hour'))
        for employee in employees:
            stt += 1
            emp = employee.sudo()
            comparedtime_from = timezone(emp.tz).localize(datetime.combine(day_start, time.min)).astimezone(timezone('UTC')).replace(tzinfo=None)
            comparedtime_to = timezone(emp.tz).localize(datetime.combine(day_end, time.max)).astimezone(timezone('UTC')).replace(tzinfo=None)
            emp_info = [
                stt,
                emp.barcode,
                emp.name,
                emp.department_id.display_name,
                '', # tình trạng hợp đồng
                emp.en_date_start.strftime('%d/%m/%Y') if emp.en_date_start else '',
                emp.date_end_probation.strftime('%d/%m/%Y') if emp.date_end_probation else '',
                '', # thay đổi lương
            ]

            # month_leave_intervals = employee.list_leaves(comparedtime_from, comparedtime_to)
            work_day = 0
            probation = 0
            old_salary = 0
            total_paid = 0
            total_leave = 0
            cd = nl_nb = nm = ts = kl = off = ds = o = 0
            total_day_month = 0
            total_day_have = 0
            for i in range(1, 32):
                if i <= max_day:
                    date_check = date(int(objs.year), int(objs.month), i)
                    comparedtime_from = timezone(emp.tz).localize(datetime.combine(date_check, time.min)).astimezone(timezone('UTC')).replace(tzinfo=None)
                    comparedtime_to = timezone(emp.tz).localize(datetime.combine(date_check, time.max)).astimezone(timezone('UTC')).replace(tzinfo=None)
                    dict_leave_day_intervals = employee.list_leaves(comparedtime_from, comparedtime_to)
                    dict_leave_day = []
                    for day, hours, leave in dict_leave_day_intervals:
                        dict_leave_day.append([leave, hours])
                    attendances = self.env['hr.attendance'].sudo().search([('employee_id', '=', emp.id), ('date', '=', date_check)]).mapped('worked_hours')
                    calendar_leaves_ts = self.env['resource.calendar.leaves'].sudo().search([('date_from_convert', '<=', date_check), ('date_to_convert', '>=', date_check), ('resource_id', '=', employee.resource_id.id), ('name', 'like', "Nghỉ thai sản")])
                    calendar_leaves_kl = self.env['resource.calendar.leaves'].sudo().search([('date_from_convert', '<=', date_check), ('date_to_convert', '>=', date_check), ('resource_id', '=', employee.resource_id.id), ('name', 'like', "Nghỉ không lương")])

                    total_hour = sum(attendances)
                    if employee.departure_date and employee.departure_date <= date_check:
                        # Nghỉ việc
                        emp_info.append('')
                    elif not employee.en_date_start or employee.en_date_start > date_check:
                        # Không có ngày bắt đầu
                        # chưa đi làm
                        emp_info.append('')
                    elif employee.en_status_hr in ['maternity-leave', 'semi-inactive'] and (not employee.en_day_layoff_from or employee.en_day_layoff_from <= date_check) and (not employee.en_day_layoff_to or employee.en_day_layoff_to >= date_check):
                        # trong khoảng bị nghỉ việc/nghỉ thai sản
                        if employee.en_status_hr == 'maternity-leave':
                            emp_info.append('TS')
                            ts += 1
                        else:
                            emp_info.append('')
                    elif employee.en_status_hr in ['active'] and calendar_leaves_ts:
                        # trong tháng có khoảng nghỉ thai sản
                        emp_info.append('TS')
                        ts += 1
                    elif employee.en_status_hr in ['active'] and calendar_leaves_kl:
                        # trong tháng có khoảng nghỉ không lương
                        emp_info.append('')
                    elif date_check.weekday() == 6:
                        # CN
                        emp_info.append('OFF')
                        off += 1
                    elif date_check.weekday() == 5:
                        # T7
                        if dict_leave_day and any(leave.is_holiday for leaves, hours in dict_leave_day for leave in leaves):
                            emp_info.append('NL')
                            nl_nb += 0.5
                        else:
                            emp_info.append('OFF/2')
                            off += 0.5
                    else:
                        # Nghỉ lễ
                        if dict_leave_day and any(leave.is_holiday for leaves, hours in dict_leave_day for leave in leaves):
                            emp_info.append('NL')
                            nl_nb += 1
                            continue
                        # Ngày thường
                        # Xét nghỉ phép trước
                        # Sau đó xét chấm công
                        # Không có chấm công thì là Ro
                        lst_leave = []
                        man_day = 0
                        mark_missing_attendance = False
                        if dict_leave_day:
                            total_by_leave = defaultdict(float)
                            for leaves, hours in dict_leave_day:
                                for leave in leaves:
                                    code = leave.holiday_id.holiday_status_id.code
                                    if not code:
                                        continue
                                        # code = 'NA'
                                    rate_time = 0.5
                                    if leave.holiday_id.use_time_slot or leave.holiday_id.request_unit_hours:
                                        rate_time = leave.holiday_id.number_of_days

                                    total_by_leave[code] += rate_time
                                    man_day += rate_time
                                    if code == 'P':
                                        total_leave += rate_time
                                    elif code == 'SN':
                                        total_leave += rate_time
                                    elif code == 'CD':
                                        cd += rate_time
                                    elif code == 'NB':
                                        nl_nb += rate_time
                                    elif code == 'NM':
                                        nm += rate_time
                                    elif code == 'TS':
                                        ts += rate_time
                                    elif code == 'Ô':
                                        cd += rate_time
                                    elif code == 'DS':
                                        cd += rate_time
                                    elif code == 'Ro':
                                        kl += rate_time
                            for code, total in total_by_leave.items():
                                if not total % 1:
                                    total = int(total)
                                if code == "NM":
                                    lst_leave.append(f"{1 if total != 1 else total}{code}")
                                else:
                                    lst_leave.append(f"{total if total != 1 else ''}{code}")
                        if not man_day:
                            if employee.shift == 'ca_vip':
                                lst_leave.append('X')
                            elif attendances:
                                hour_per_day = 8
                                if employee.has_child_start and employee.has_child_start <= date_check and (not employee.has_child_end or employee.has_child_end >= date_check):
                                    hour_per_day = 7
                                if total_hour >= (hour_per_day - 0.25):
                                    lst_leave.append('X')
                                elif total_hour >= 0.75 * hour_per_day:
                                    lst_leave.append('0.75X')
                                    mark_missing_attendance = True
                                elif total_hour >= 0.4375 * hour_per_day:
                                    lst_leave.append('0.5X')
                                    mark_missing_attendance = True
                                elif total_hour >= 0.25 * hour_per_day:
                                    lst_leave.append('0.25X')
                                    mark_missing_attendance = True
                                else:
                                    lst_leave.append('0X')
                                    mark_missing_attendance = True
                            else:
                                lst_leave.append('Ro')
                                mark_missing_attendance = True
                                kl += 1
                        elif man_day < 1:
                            if employee.shift == 'ca_vip':
                                lst_leave.append(f"{1 - man_day}X")
                            elif attendances:
                                attendance_man_day = 0
                                hour_per_day = 8
                                if employee.has_child_start and employee.has_child_start <= date_check and (not employee.has_child_end or employee.has_child_end >= date_check):
                                    hour_per_day = 7
                                if total_hour >= (hour_per_day - 0.25):
                                    attendance_man_day = 1
                                elif total_hour >= 0.75 * hour_per_day:
                                    attendance_man_day = 0.75
                                elif total_hour >= 0.4375 * hour_per_day:
                                    attendance_man_day = 0.5
                                elif total_hour >= 0.25 * hour_per_day:
                                    attendance_man_day = 0.25
                                else:
                                    mark_missing_attendance = True
                                if (1 - man_day) > attendance_man_day:
                                    mark_missing_attendance = True
                                    lst_leave.append(f"{attendance_man_day}X")
                            else:
                                ro_day = 1 - man_day
                                if '0.5Ro' in lst_leave:
                                    lst_leave.remove('0.5Ro')
                                    ro_day += 0.5
                                lst_leave.append(f"{ro_day if ro_day != 1 else ''}Ro")
                                kl += 1 - man_day
                        message_leave = '/'.join(lst_leave)
                        if employee.check_timesheet_before_checkout and -1 < employee.get_hour_working_by_day(date_check) < check_hour:
                            message_leave += '-'
                            mark_missing_attendance = True
                        if mark_missing_attendance:
                            message_leave += 'mark_missing_attendance'
                        emp_info.append(message_leave)
                else:
                    emp_info.append('')
            # work_day = emp._get_work_days_data(
            #     day_start,
            #     day_end,
            #     calendar=emp.resource_calendar_id,
            #     compute_leaves=False,
            # )['days']
            emp_info += [
                work_day,
                probation,
                old_salary,
                total_paid,
                total_leave,
                cd,
                nl_nb,
                nm,
                ts,
                kl,
                off,
                total_day_month,
                total_day_have
            ]
            for col, val in enumerate(emp_info):
                write_val = val
                if type(val) == str and 'mark_missing_attendance' in val:
                    write_val = val.replace('mark_missing_attendance', '')
                    sheet.cell(row_to_write, col + 1).fill = redFill
                sheet.cell(row_to_write, col+1).value = write_val or ''
            row_to_write += 1
        return wb

    def _export_by_custom(self, wb, sheet, objs):
        """Tạo báo cáo theo khoảng ngày tùy chọn"""
        row_to_write = 7
        stt = 0
        day_start = objs.date_from
        day_end = objs.date_to
        sheet['A2'] = ''
        sheet['A3'] = ''
        sheet['N3'] = 'Từ %s ~ %s' % (day_start.strftime("%d/%m/%Y"), day_end.strftime("%d/%m/%Y"))
        # Thiết lập các tiêu đề cột theo khoảng ngày chọn
        day_of_week_map = {0: 'T2', 1: 'T3', 2: 'T4', 3: 'T5', 4: 'T6', 5: 'T7', 6: 'CN'}
        day_in_month_col = 9
        for i in range((day_end - day_start).days + 1):
            current_date = day_start + timedelta(days=i)  # Tính ngày hiện tại
            sheet.cell(row_to_write - 2, day_in_month_col + i, day_of_week_map.get(current_date.weekday()))
            sheet.cell(row_to_write - 1, day_in_month_col + i, current_date.day)
        employees = objs.with_context(active_test=False).employee_ids
        if not employees:
            domain = ['|', ('active', '=', True),
                      '&',
                      ('active', '=', False),
                      '|',
                      '&',
                      '|', ('en_day_layoff_from', '<=', day_end), ('en_day_layoff_from', '=', False),
                      '|', ('en_day_layoff_to', '>=', day_start), ('en_day_layoff_to', '=', False),
                      ('departure_date', '>', day_start)
                      ]
            if objs.department_ids:
                domain += [('department_id', 'in', objs.department_ids.ids)]
            if objs.en_department_ids:
                domain += [('en_department_id', 'in', objs.en_department_ids.ids)]
            employees = self.env['hr.employee'].search(domain)
        check_hour = float(self.env['ir.config_parameter'].sudo().get_param('check_timesheet_before_checkout_hour'))
        for employee in employees:
            day_start = objs.date_from
            day_end = objs.date_to
            stt += 1
            emp = employee.sudo()
            emp_info = [stt, emp.barcode, emp.name, emp.department_id.display_name, '',  # tình trạng hợp đồng
                emp.en_date_start.strftime('%d/%m/%Y') if emp.en_date_start else '',
                emp.date_end_probation.strftime('%d/%m/%Y') if emp.date_end_probation else '', '',  # thay đổi lương
            ]
            work_day = 0
            probation = 0
            old_salary = 0
            total_paid = 0
            total_leave = 0
            cd = nl_nb = nm = ts = kl = off = ds = o = 0
            total_day_month = 0
            total_day_have = 0
            for i in range(0, 32):
                date_check = day_start
                if i > 30:
                    continue
                if day_start > day_end:
                    emp_info.append('')
                    day_start += timedelta(days=1)
                    continue
                day_start += timedelta(days=1)
                comparedtime_from = timezone(emp.tz).localize(datetime.combine(date_check, time.min)).astimezone(
                    timezone('UTC')).replace(tzinfo=None)
                comparedtime_to = timezone(emp.tz).localize(datetime.combine(date_check, time.max)).astimezone(
                    timezone('UTC')).replace(tzinfo=None)
                dict_leave_day_intervals = employee.list_leaves(comparedtime_from, comparedtime_to)
                dict_leave_day = []
                for day, hours, leave in dict_leave_day_intervals:
                    dict_leave_day.append([leave, hours])
                attendances = self.env['hr.attendance'].sudo().search([('employee_id', '=', emp.id), ('date', '=', date_check)]).mapped('worked_hours')
                calendar_leaves_ts = self.env['resource.calendar.leaves'].sudo().search([('date_from_convert', '<=', date_check), ('date_to_convert', '>=', date_check), ('resource_id', '=', employee.resource_id.id), ('name', 'like', "Nghỉ thai sản")])
                calendar_leaves_kl = self.env['resource.calendar.leaves'].sudo().search([('date_from_convert', '<=', date_check), ('date_to_convert', '>=', date_check), ('resource_id', '=', employee.resource_id.id), ('name', 'like', "Nghỉ không lương")])

                total_hour = sum(attendances)
                if employee.departure_date and employee.departure_date <= date_check:
                    # Nghỉ việc
                    emp_info.append('')
                elif not employee.en_date_start or employee.en_date_start > date_check:
                    # Không có ngày bắt đầu
                    # chưa đi làm
                    emp_info.append('')
                elif employee.en_status_hr in ['maternity-leave', 'semi-inactive'] and (
                        not employee.en_day_layoff_from or employee.en_day_layoff_from <= date_check) and (
                        not employee.en_day_layoff_to or employee.en_day_layoff_to >= date_check):
                    # trong khoảng bị nghỉ việc/nghỉ thai sản
                    if employee.en_status_hr == 'maternity-leave':
                        emp_info.append('TS')
                        ts += 1
                    else:
                        emp_info.append('')
                elif employee.en_status_hr in ['active'] and calendar_leaves_ts:
                    # trong tháng có khoảng nghỉ thai sản
                    emp_info.append('TS')
                    ts += 1
                elif employee.en_status_hr in ['active'] and calendar_leaves_kl:
                    # trong tháng có khoảng nghỉ không lương
                    emp_info.append('')
                elif date_check.weekday() == 6:
                    # CN
                    emp_info.append('OFF')
                    off += 1
                elif date_check.weekday() == 5:
                    # T7
                    if dict_leave_day and any(
                            leave.is_holiday for leaves, hours in dict_leave_day for leave in leaves):
                        emp_info.append('NL')
                        nl_nb += 0.5
                    else:
                        emp_info.append('OFF/2')
                        off += 0.5
                else:
                    # Nghỉ lễ
                    if dict_leave_day and any(
                            leave.is_holiday for leaves, hours in dict_leave_day for leave in leaves):
                        emp_info.append('NL')
                        nl_nb += 1
                        continue
                    # Ngày thường
                    # Xét nghỉ phép trước
                    # Sau đó xét chấm công
                    # Không có chấm công thì là Ro
                    lst_leave = []
                    man_day = 0
                    mark_missing_attendance = False
                    if dict_leave_day:
                        total_by_leave = defaultdict(float)
                        for leaves, hours in dict_leave_day:
                            for leave in leaves:
                                code = leave.holiday_id.holiday_status_id.code
                                if not code:
                                    continue
                                    # code = 'NA'
                                rate_time = 0.5
                                if leave.holiday_id.use_time_slot or leave.holiday_id.request_unit_hours:
                                    rate_time = leave.holiday_id.number_of_days

                                total_by_leave[code] += rate_time
                                man_day += rate_time
                                if code == 'P':
                                    total_leave += rate_time
                                elif code == 'SN':
                                    total_leave += rate_time
                                elif code == 'CD':
                                    cd += rate_time
                                elif code == 'NB':
                                    nl_nb += rate_time
                                elif code == 'NM':
                                    nm += rate_time
                                elif code == 'TS':
                                    ts += rate_time
                                elif code == 'Ô':
                                    cd += rate_time
                                elif code == 'DS':
                                    cd += rate_time
                                elif code == 'Ro':
                                    kl += rate_time
                        for code, total in total_by_leave.items():
                            if not total % 1:
                                total = int(total)
                            if code == "NM":
                                lst_leave.append(f"{1 if total != 1 else total}{code}")
                            else:
                                lst_leave.append(f"{total if total != 1 else ''}{code}")
                    if not man_day:
                        if employee.shift == 'ca_vip':
                            lst_leave.append('X')
                        elif attendances:
                            hour_per_day = 8
                            if employee.has_child_start and employee.has_child_start <= date_check and (
                                    not employee.has_child_end or employee.has_child_end >= date_check):
                                hour_per_day = 7
                            if total_hour >= (hour_per_day - 0.25):
                                lst_leave.append('X')
                            elif total_hour >= 0.75 * hour_per_day:
                                lst_leave.append('0.75X')
                                mark_missing_attendance = True
                            elif total_hour >= 0.4375 * hour_per_day:
                                lst_leave.append('0.5X')
                                mark_missing_attendance = True
                            elif total_hour >= 0.25 * hour_per_day:
                                lst_leave.append('0.25X')
                                mark_missing_attendance = True
                            else:
                                lst_leave.append('0X')
                                mark_missing_attendance = True
                        else:
                            lst_leave.append('Ro')
                            mark_missing_attendance = True
                            kl += 1
                    elif man_day < 1:
                        if employee.shift == 'ca_vip':
                            lst_leave.append(f"{1 - man_day}X")
                        elif attendances:
                            attendance_man_day = 0
                            hour_per_day = 8
                            if employee.has_child_start and employee.has_child_start <= date_check and (
                                    not employee.has_child_end or employee.has_child_end >= date_check):
                                hour_per_day = 7
                            if total_hour >= (hour_per_day - 0.25):
                                attendance_man_day = 1
                            elif total_hour >= 0.75 * hour_per_day:
                                attendance_man_day = 0.75
                            elif total_hour >= 0.4375 * hour_per_day:
                                attendance_man_day = 0.5
                            elif total_hour >= 0.25 * hour_per_day:
                                attendance_man_day = 0.25
                            else:
                                mark_missing_attendance = True
                            if (1 - man_day) > attendance_man_day:
                                mark_missing_attendance = True
                                lst_leave.append(f"{attendance_man_day}X")
                        else:
                            ro_day = 1 - man_day
                            if '0.5Ro' in lst_leave:
                                lst_leave.remove('0.5Ro')
                                ro_day += 0.5
                            lst_leave.append(f"{ro_day if ro_day != 1 else ''}Ro")
                            kl += 1 - man_day
                    message_leave = '/'.join(lst_leave)
                    if employee.check_timesheet_before_checkout and -1 < employee.get_hour_working_by_day(
                            date_check) < check_hour:
                        message_leave += '-'
                        mark_missing_attendance = True
                    if mark_missing_attendance:
                        message_leave += 'mark_missing_attendance'
                    emp_info.append(message_leave)
            emp_info += [work_day, probation, old_salary, total_paid, total_leave,
                cd, nl_nb, nm, ts, kl, off, total_day_month, total_day_have]
            for col, val in enumerate(emp_info):
                write_val = val
                if type(val) == str and 'mark_missing_attendance' in val:
                    write_val = val.replace('mark_missing_attendance', '')
                    sheet.cell(row_to_write, col + 1).fill = redFill
                sheet.cell(row_to_write, col + 1).value = write_val or ''
            row_to_write += 1
        return wb
