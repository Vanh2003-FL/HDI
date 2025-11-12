from odoo import models, _,fields
import pathlib
from odoo.exceptions import ValidationError
from openpyxl import load_workbook
from odoo.addons.ngsd_report_xlsx.utils import xlsx_utils
from datetime import timedelta, datetime, date


class LeadExportReportXlsx(models.AbstractModel):
    _name = 'report.ngsd_base.report_lead'
    _inherit = 'report.report_xlsx.abstract'

    def generate_xlsx_report(self, workbook, data, objs):
        template_name = 'lead_report.xlsx'
        template_file = pathlib.Path(__file__).parent.parent.joinpath('static', 'excel', template_name)
        try:
            wb = load_workbook(template_file)
        except FileNotFoundError:
            raise ValidationError(_("Error: Could not find template {}").format(template_name))

        sheet = wb['Mẫu báo cáo']
        row_to_write = 4
        objs.export_data(['line_ids'])
        stt = 0
        for obj in objs:
            stt += 1
            obj = obj.sudo()
            for idx, line in enumerate(obj.line_ids):
                data_line = [
                    stt,
                    obj.create_date.strftime('%d/%m/%Y') if obj.create_date else '',
                    obj.partner_id.display_name,
                    obj.code,
                    obj.contract_code,
                    obj.display_name,
                    'Mức độ ' + (obj.priority or '0'),
                    line.product_name or '',
                    line.partner_id.display_name,
                    line.estimated_value,
                    obj.tender_model_id.display_name,
                    obj.role_id.display_name,
                    line.ngsd_revenue,
                    line.ngsd_revenue_taxed,
                    obj.budget_approval_date.strftime('%d/%m/%Y') if obj.budget_approval_date else '',
                    obj.bid_opening_date.strftime('%d/%m/%Y') if obj.bid_opening_date else '',
                    obj.bid_closing_date.strftime('%d/%m/%Y') if obj.bid_closing_date else '',
                    obj.date_deadline.strftime('%d/%m/%Y') if obj.date_deadline else '',
                    obj.kickoff_planned_date.strftime('%d/%m/%Y') if obj.kickoff_planned_date else '',
                    obj.project_end_date.strftime('%d/%m/%Y') if obj.project_end_date else '',
                    obj.stage_id.display_name,
                    ', '.join(obj.solution_architect_uids.mapped('display_name')),
                    obj.user_id.display_name,
                    obj.team_id.display_name,
                ]
                for col, val in enumerate(data_line):
                    sheet.cell(row_to_write, col+1).value = val or ''
                row_to_write += 1

        return wb
