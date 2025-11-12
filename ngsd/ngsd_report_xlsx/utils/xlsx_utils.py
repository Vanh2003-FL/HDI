from copy import copy
import re
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import range_boundaries
import logging
_logger = logging.getLogger(__name__)


def float_2_commas(number):
    try:
        sub = ''
        str_num = str(int(float(abs(number))))
        if number < 0:
            sub = '-'
        return sub + ''.join(reversed([x + ('.' if i and not i % 3 else '') for i, x in enumerate(reversed(str_num))]))
    except:
        return number

def find_text(sheet, text, type):
    row_ = None
    for row in sheet.rows:
        for cell in row:
            if cell.value == text:
                if type == 'row':
                    row_ = cell.row
                if type == 'cell':
                    row_ = cell.column_letter
                if type == 'both':
                    row_ = cell.coordinate
    return row_

def get_merged_cells(sheet, row):
    cells = []
    if row:
        for mcr in sheet.merged_cells:
            if str(row) in re.findall(r'\d+', str(mcr)):
                cells.append(re.sub(r'\d+', '{row}', str(mcr)))
    return cells

def merge_cells(sheet, row, merged_cells):
    for cell in merged_cells:
        sheet.merge_cells(cell.format(row=row))

def set_style_cells(sheet, cell, style):
    sheet[cell]._style = copy(style)

def write_value(sheet, row, cell, value):
    if row:
        row_start_srt = str(row)
        sheet[cell + row_start_srt] = value
    else:
        sheet[cell] = value

def write_value_to_excel(sheet, row, cell, value, copy=False, location='top'):
    write_value(sheet, row, cell, value)
    if copy:
        copy_style_cells(sheet, row, location)
    # merged_cells_list = get_merged_cells(sheet, row)
    # if merge_cell:
    #     merge_cells(sheet, row, merged_cells_list)

def move_rows(sheet, row, number=1):
    dimensions = re.sub(r'\d+', '{row}', sheet.dimensions)
    list_range = list(range(row, sheet.max_row + 1))
    list_range.sort(reverse=True)
    for r in list_range:
        ran = dimensions.format(row=r)
        move_merged_cells(sheet, r, number)
        sheet.move_range(ran, number)

def unmerge_cells(sheet, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
    """ Remove merge on a cell range.  Range is a cell range (e.g. A1:E1) """
    cr = CellRange(range_string=range_string, min_col=start_column, min_row=start_row,
                  max_col=end_column, max_row=end_row)

    if cr.coord not in sheet.merged_cells:
        raise ValueError("Cell range {0} is not merged".format(cr.coord))

    sheet.merged_cells.remove(cr)

def unmerge_footer(sheet, rows):
    affected_cells = []
    for row in rows:
        for cell in sheet.merged_cells:
            merged_cell_start = re.sub(r':+[A-Z]{1,}[0-9]{1,}', '', cell.coord)
            merged_row_start = re.sub(r'[A-Z]{1,}', '', merged_cell_start)
            if row == int(merged_row_start):
                affected_cells.append(cell.coord)
    for cell in affected_cells:
        cell_range = range_boundaries(cell)
        unmerge_cells(sheet,
                      start_column=cell_range[0],
                      start_row=cell_range[1],
                      end_column=cell_range[2],
                      end_row=cell_range[3]
                      )

def copy_style_cells(sheet, row, location='top'):
    if location == 'bottom':
        row_style = row + 1
    elif location == 'top':
        row_style = row - 1
    else:
        row_style = location
    for cel in sheet[row_style]:
        sheet.cell(row=row, column=cel.column)._style = copy(sheet.cell(row=row_style, column=cel.column)._style)

def copy_style_cells_to_sheet(sheet_copy, sheet_to, row, location='top'):
    if location == 'bottom':
        row_style = row + 1
    elif location == 'top':
        row_style = row - 1
    else:
        row_style = location
    for cel in sheet_copy[row_style]:
        sheet_to.cell(row=row, column=cel.column)._style = copy(sheet_copy.cell(row=row_style, column=cel.column)._style)

def clear_style_cells(sheet, row):
    for cel in sheet[row]:
        sheet.cell(row=row, column=cel.column)._style = None

def check_merged_cell(cell):
    if 'MergedCell' in str(type(cell)):
        return True
    return False

def move_merged_cells(sheet, row, number=1, on_row=False, copy_style=True):
    affected_cells = []
    cells_value = []
    coord_merged = []
    # lấy ra các ô được merged
    for cell in sheet.merged_cells:
        merged_cell_start = re.sub(r':+[A-Z]{1,}[0-9]{1,}', '', cell.coord)
        merged_row_start = re.sub(r'[A-Z]{1,}', '', merged_cell_start)
        if row == int(merged_row_start):
            affected_cells.append(cell.coord)
            cells_value.append(re.sub(r':+[A-Z]{1,}[0-9]{1,}', '', cell.coord))
        coord_merged.append(cell.coord) # lấy danh sách các coord được merge
    # duyệt từng ô được merged và unmerge nó
    for cell in affected_cells:
        cell_range = range_boundaries(cell)
        unmerge_cells(sheet,
                      start_column=cell_range[0],
                      start_row=cell_range[1],
                      end_column=cell_range[2],
                      end_row=cell_range[3]
                      )
    if copy_style:
        copy_style_cells(sheet, row + number, row) # copy style ô cũ và set cho ô merged mới
    # duyệt đến danh sách ô merged và tăng giá trị theo number. sau đó merge ô mới
    for cell in affected_cells:
        cell_range = range_boundaries(cell)
        start_column = cell_range[0]
        end_column = cell_range[2]
        start_row = cell_range[1] + number
        end_row = cell_range[3] + number
        if on_row: # trường hợp di chuyển ô merged trong cùng 1 row
            start_column = cell_range[0] + number
            end_column = cell_range[2] + number
            start_row = cell_range[1]
            end_row = cell_range[3]
        sheet.merge_cells(start_column=start_column, start_row=start_row, end_column=end_column, end_row=end_row)
        for cel in cells_value: # set giá trị cho ô mới
            if sheet[cel].value:
                sheet.cell(start_row, start_column).value = sheet[cel].value
    for cell in sheet[row]:
        # kiểm tra các ô nằm trong các ô được merge thì bỏ qua
        if 'MergedCell' in str(type(cell)):
            continue
        coordinate = re.sub(r'\d+', str(row + number), cell.coordinate)
        sheet[coordinate].value = sheet[cell.coordinate].value

def render_fixed_datas(sheet, datas):
    for row in sheet.rows:
        for cell in row:
            if cell.value in datas:
                try:
                    cell.value = datas.get(cell.value, '') or ''
                except:
                    continue

def render_loop_data(sheet, datas):
    # def print_loop_datas(sheet, row_start=None, cells=[], values={}):
    #     while values:
    #         keys = list(values.keys())
    #         # cells = cells
    #         for j in keys:
    #             if type(values[j]) == dict:
    #                 child_keys = list(values[j].keys())
    #                 for i in child_keys:
    #                     target = find_text(sheet, j + '.' + i, 'both')
    #                     cells.append(target)
    #                     if not row_start:
    #                         row_start = re.sub(r'[A-Z]', '', target)
    #                         row_start = int(row_start) if row_start else None
    #                         # move_rows(sheet, row_start + 1, 1)
    #                         sheet[target].value = values[j].get(i, '')
    #                     else:
    #                         if target:
    #                             sheet[target].value = values[j].get(i, '')
    #                     if type(values[j][i]) == dict:
    #
    #                 print_loop_datas(sheet, row_start, values[j])
    #             del values[j]
    # print_loop_datas(sheet, None, [], datas)

    def print_loop_datas(sheet, row_start=None, cells={}, var='', values={}):
        while values:
            keys = list(values.keys())
            for j in keys:
                if not row_start:
                    if type(values[j]) == dict:
                        var = re.sub(r'_.+', '', j)
                        cells[var] = []
                        child_keys = list(values[j].keys())
                        for i in child_keys:
                            target = find_text(sheet, var + '.' + i, 'both')
                            if not target:
                                continue
                            # sheet[target].value = values[j].get(i, '')
                            if not cells:
                                if target not in cells[var]:
                                    cells[var].append(target)

                        if not row_start:
                            row_start = re.sub(r'[A-Z]', '', cells[0])
                            row_start = int(row_start) if row_start else None
                            # move_rows(sheet, row_start + 1, 1)
                        print_loop_datas(sheet, row_start, cells, var, values[j])
                    del values[j]
                else:
                    if not type(values[j]) == dict:
                        target = cells.get(j, '')
                        if not target:
                            continue
                        sheet[target].value = values.get(j, '')
                    else:
                        pass
                        # var = re.sub(r'_.+', '', j)
                        # cells = {}
                        # cells[var] = []
                        # child_keys = list(values[j].keys())
                        # for i in child_keys:
                        #     target = find_text(sheet, var + '.' + i, 'both')
                        #     if not target:
                        #         continue
                        #     sheet[target].value = values[j].get(i, '')
                        #     if not cells:
                        #         if target not in cells[var]:
                        #             cells[var].append(target)


    print_loop_datas(sheet, datas)

def supreme_render(sheet, datas_is_fixed, datas_is_loop):
    render_fixed_datas(sheet, datas_is_fixed)

    render_loop_data(sheet, datas_is_loop)


def render_to_excel(sheet, row, values, dict_style={}, dict_write_value={}, merged_cells=[], var_count_up='', count=0):
    for cell in dict_write_value.keys():
        if dict_write_value[cell] == var_count_up:
            count += 1
            value = str(count)
        else:
            value = values[dict_write_value[cell]]
        write_value(sheet, row, cell, value)
        set_style_cells(sheet, cell + str(row), dict_style[cell])
    merge_cells(sheet, row, merged_cells)

def get_write_and_style_row(sheet, row, keys):
    dict_write = {}
    dict_style = {}
    for cell in sheet[row]:
        if cell.value in keys:
            var = cell.value[cell.value.find('.') + 1:]
            dict_write[cell.column_letter] = var
        if check_merged_cell(cell) and cell.value not in keys:
            continue
        dict_style[cell.column_letter] = sheet[cell.coordinate]._style
    return dict_write, dict_style

def get_merged_rows(sheet, rows):
    affected_cells = []
    for row in rows:
        for cell in sheet.merged_cells:
            merged_cell_start = re.sub(r':+[A-Z]{1,}[0-9]{1,}', '', cell.coord)
            merged_row_start = re.sub(r'[A-Z]{1,}', '', merged_cell_start)
            if row == int(merged_row_start):
                affected_cells.append(cell.coord)
    return affected_cells


def get_write_value_style_row(sheet, row):
    dict_write = {}
    dict_style = {}
    dict_value = {}
    for cell in sheet[row]:
        try:
            column_letter = cell.column_letter
        except:
            continue
        dict_write[column_letter] = column_letter
        dict_value[column_letter] = cell.value
        dict_style[column_letter] = sheet[cell.coordinate]._style
    return dict_write, dict_style, dict_value

def copy_row(sheet=False, from_row=0, to_row=0, style=True):
        """
        :param from_row: dòng cần copy
        :param to_row: dòng cần in
        :param to_row: dòng cần in
        :param style: có cần lấy style hay không
        """
        if not sheet or not int(from_row) or not int(to_row):
            return
        sheet.move_range("A{0}:L{0}".format(from_row), rows=to_row - from_row, cols=0)
        for mcr in sheet.merged_cells:
            if str(from_row) in re.findall(r'\d+', str(mcr)):
                string_row = str(mcr).replace(str(from_row), str(to_row))
                sheet.merge_cells(string_row)
        if style:
            for cell in sheet[from_row]:
                new_cell = sheet.cell(row=to_row, column=cell.column)
                if new_cell.has_style:
                    cell.font = copy(new_cell.font)
                    cell.border = copy(new_cell.border)
                    cell.fill = copy(new_cell.fill)
                    cell.number_format = copy(new_cell.number_format)
                    cell.protection = copy(new_cell.protection)
                    cell.alignment = copy(new_cell.alignment)


def merge_footer(sheet, row, cells):
    for i in cells:
        row_cell = re.findall(r'\d+', i)
        cell_range_str = re.findall(r'[A-Z]{1,}', i)
        j = int(row_cell[1]) - int(row_cell[1])
        cell_range = cell_range_str[0] + str(row) + ':' + cell_range_str[1] + str(row)
        if j > 0:
            cell_range = cell_range_str[0] + str(row +j) + ':' + cell_range_str[1] + str(row)
        cell_range_tuple = range_boundaries(cell_range)
        start_column = cell_range_tuple[0]
        end_column = cell_range_tuple[2]
        start_row = cell_range_tuple[1]
        end_row = cell_range_tuple[3]
        sheet.merge_cells(start_column=start_column, start_row=start_row, end_column=end_column, end_row=end_row)

def render_to_excel(sheet, row, values, dict_style={}, dict_write_value={}, merged_cells=[], var_count_up='', count=0):
    for cell in dict_write_value.keys():
        if dict_write_value[cell] == var_count_up:
            count += 1
            value = str(count)
        else:
            value = values[dict_write_value[cell]]
        write_value(sheet, row, cell, value)
        set_style_cells(sheet, cell + str(row), dict_style[cell])
    merge_cells(sheet, row, merged_cells)

def get_write_and_style_row(sheet, row, keys):
    dict_write = {}
    dict_style = {}
    for cell in sheet[row]:
        if cell.value in keys:
            var = cell.value[cell.value.find('.') + 1:]
            dict_write[cell.column_letter] = var
        if check_merged_cell(cell) and cell.value not in keys:
            continue
        dict_style[cell.column_letter] = sheet[cell.coordinate]._style
    return dict_write, dict_style

def get_write_value_style_row(sheet, row):
    dict_write = {}
    dict_style = {}
    dict_value = {}
    for cell in sheet[row]:
        try:
            column_letter = cell.column_letter
        except:
            continue
        dict_write[column_letter] = column_letter
        dict_value[column_letter] = cell.value
        dict_style[column_letter] = sheet[cell.coordinate]._style
    return dict_write, dict_style, dict_value

def copy_row(sheet=False, from_row=0, to_row=0, style=True):
        """
        :param from_row: dòng cần copy
        :param to_row: dòng cần in
        :param to_row: dòng cần in
        :param style: có cần lấy style hay không
        """
        if not sheet or not int(from_row) or not int(to_row):
            return
        sheet.move_range("A{0}:Z{0}".format(from_row), rows=to_row - from_row, cols=0)
        for mcr in sheet.merged_cells:
            if str(from_row) in re.findall(r'\d+', str(mcr)):
                string_row = str(mcr).replace(str(from_row), str(to_row))
                sheet.merge_cells(string_row)
        if style:
            for cell in sheet[from_row]:
                new_cell = sheet.cell(row=to_row, column=cell.column)
                if new_cell.has_style:
                    cell.font = copy(new_cell.font)
                    cell.border = copy(new_cell.border)
                    cell.fill = copy(new_cell.fill)
                    cell.number_format = copy(new_cell.number_format)
                    cell.protection = copy(new_cell.protection)
                    cell.alignment = copy(new_cell.alignment)


def merge_footer(sheet, row, cells):
    for i in cells:
        row_cell = re.findall(r'\d+', i)
        cell_range_str = re.findall(r'[A-Z]{1,}', i)
        j = int(row_cell[1]) - int(row_cell[1])
        cell_range = cell_range_str[0] + str(row) + ':' + cell_range_str[1] + str(row)
        if j > 0:
            cell_range = cell_range_str[0] + str(row +j) + ':' + cell_range_str[1] + str(row)
        cell_range_tuple = range_boundaries(cell_range)
        start_column = cell_range_tuple[0]
        end_column = cell_range_tuple[2]
        start_row = cell_range_tuple[1]
        end_row = cell_range_tuple[3]
        sheet.merge_cells(start_column=start_column, start_row=start_row, end_column=end_column, end_row=end_row)

####################################
# Target: copy footer excel
# By: Linh Văn

def get_values_style_multil_line(sheet, row, row_end, get_value=True):
    dict_value = {}
    dict_style = {}
    rows = list(range(row, row_end + 1))
    for row_ in rows:
        d_row_style = {}
        d_row_value = {}
        for cell in sheet[row_]:
            try:
                column_letter = cell.column_letter
                d_row_style[column_letter] = sheet[cell.coordinate]._style
                if get_value:
                    d_row_value[column_letter] = cell.value
            except:
                cell_ = re.sub(r'\d+', '', cell.coordinate)
                d_row_style[cell_] = sheet[cell.coordinate]._style
                continue
        dict_style[str(row_)] = d_row_style
        if get_value:
            dict_value[str(row_)] = d_row_value
    merged_cells = get_merged_cells(sheet, row)
    return dict_value, dict_style, merged_cells

def copy_datas_to_memory(sheet, key_word='', row_end=0, row_footer=0,):
    row_start = False
    for r in sheet.rows:
        for cell in r:
            if str(cell.value) == str(key_word):
                row_start = cell.row
                break
        if row_start:
            break
    datas = {}
    rows = list(range(row_start, row_end + 1))
    for row_ in rows:
        row_values = {}
        row_style = {}
        for cell in sheet[row_]:
            try:
                column_letter = cell.column_letter
                row_style[column_letter] = sheet[cell.coordinate]._style
                row_values[column_letter] = cell.value
            except:
                cell_ = re.sub(r'\d+', '', cell.coordinate)
                row_style[cell_] = sheet[cell.coordinate]._style
                continue
        merged_cells = get_merged_rows(sheet, [row_])
        if row_ < row_footer:
            merged_cells = get_merged_cells(sheet, row_)
        datas[str(row_)] = dict(
                                values=row_values,
                                styles=row_style,
                                merged=merged_cells,
                                )
    return datas


def merge_multil_line(sheet, row, row_to=0, merged_cells=[]):
    cells_to_merge = []
    for cell in merged_cells:
        try:
            rows_ = re.findall(r'\d+', cell)
            row_ = int(rows_[0]) + (row_to - row)
            cell_ = int(rows_[1]) + (row_to - row)
            sub_merge_cell = re.sub(r'[0-9]{1,}[:]', '{row}:', cell)
            merge_cell = re.sub(r'\d+', '{cell}', sub_merge_cell)
            cells_to_merge.append(merge_cell.format(row=row_, cell=cell_))
        except:
            continue
    merge_cells(sheet, row_to, cells_to_merge)


def render_multil_cell_on_row(sheet, row, values={}, style={}):
    for cell in sheet[row]:
        try:
            cell_ = cell.column_letter
        except:
            cell_ = re.sub(r'\d+', '', cell.coordinate)
        if cell_ in values:
            cell.value = values[cell_]
        set_style_cells(sheet, cell_ + str(row), style[cell_])


def render_multil_line(sheet, row_to, values={}):
    row_start = False
    for row in values:
        if not row_start:
            row_start = int(row)
        row_ = int(row) + (row_to - row_start)
        render_multil_cell_on_row(sheet, row_, values[row]['values'], values[row]['styles'])


def get_cell_and_style_1_row_with_keys(sheet, row, keys=[]):
    cells = {}
    styles = {}
    for cell in sheet[row]:
        cell_ = re.sub(r'\d+', '', cell.coordinate)
        styles[cell_] = cell._style
        if keys and cell.value in keys:
            row_ = cell.column_letter
            cells[cell.value] = row_
    return cells, styles


# def get_cells_with_keys(sheet, keys=[]):
#     """
#     Hàm này sẽ:
#     - Kiểm tra các key được truyền vào,
#      nếu thấy thì gán key đấy với giá trị là cột được tìm thấy
#     - Lấy tất cả style các ô của dòng được tìm thấy
#     - Lấy tất cả các ô đã merge của dòng được tìm thấy
#     :param sheet:
#     :param keys:
#     :return:
#     """
#     cells = {}
#     styles = {}
#     row = False
#     for r in sheet.rows:
#         for cell in r:
#             if cell.value in keys:
#                 row = cell.row
#                 break
#         if row:
#             cells, styles = get_cell_and_style_1_row_with_keys(sheet, row, keys)
#             break
#     merged_cells = get_merged_cells(sheet, row)
#     return cells, styles, merged_cells

def pretty_render_to_excel(sheet, row, cells, styles, merged_cells, values={}, sequence='', count=0, default=''):
    """
    :param sheet:
    :param row: dòng sẽ viết
    :param cells: dict cột theo key
    :param styles: dict style theo cột
    :param merged_cells: list ô đã merge của dòng
    :param values: dict value
    :param sequence: ô sẽ tự động tăng sequence
    :param count: bắt đầu tăng sequence từ count
    :return:
    """
    for cell in cells:
        keyword = cells.get(cell, False)
        value = values.get(keyword, '') or default
        if keyword:
            if sequence and sequence == keyword:
                count += 1
                value = str(count)
            sheet[cell + str(row)].value = value
        set_style_cells(sheet, cell + str(row), styles[cell])
    merge_cells(sheet, row, merged_cells)

def clean_demo_lines_and_footer_datas(sheet, row_start, row_end):
    """
    :param sheet:
    :param row_start: dòng bắt đầu xuất
    :param row_end: dòng tối đa của sheet mẫu
    :return:
    """
    unmerge_footer(sheet, list(range(row_start, row_end + 1)))
    for i in range(row_start, row_end + 1):
        sheet.delete_rows(row_start)


def render_footer_datas_again(sheet, row_start, row_footer_start, row_footer_max, datas):
    """
    :param sheet:
    :param row_start: dòng bắt đầu xuất
    :param row_footer_start: dòng bắt đầu của footer
    :param merged_lines_footer: danh sách dòng đã merge ở footer
    :param dict_value: giá trị (dict)
    :param dict_style: style (dict)
    :return:
    """
    merged_lines_footer = []
    footer_datas = {}
    for i in range(row_footer_start, row_footer_max+1):
        merged_lines_footer += datas[str(i)]['merged']
        footer_datas[str(i)] = datas[str(i)]
    render_multil_line(sheet, row_start, footer_datas)
    merge_multil_line(sheet, row_footer_start, row_start, merged_lines_footer)